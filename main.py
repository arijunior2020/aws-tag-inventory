import boto3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, PieChart

def get_tagged_resources(profile_name="finops-tag-inventory"):
    session = boto3.Session(profile_name=profile_name)
    client = session.client('resourcegroupstaggingapi')
    resources = []
    paginator = client.get_paginator('get_resources')
    for page in paginator.paginate(ResourcesPerPage=50):
        for res in page['ResourceTagMappingList']:
            resources.append({
                'ResourceARN': res['ResourceARN'],
                'ResourceType': res['ResourceARN'].split(":")[2],
                'Region': res['ResourceARN'].split(":")[3],
                'Tags': {tag['Key']: tag['Value'] for tag in res.get('Tags', [])}
            })
    return pd.DataFrame(resources)

def export_to_excel(df, output_path):
    df['TagsStr'] = df['Tags'].apply(lambda t: ", ".join(f"{k}={v}" for k, v in t.items()))

    # Estatísticas
    type_count = df['ResourceType'].value_counts().reset_index()
    type_count.columns = ['ResourceType', 'Count']

    region_count = df['Region'].value_counts().reset_index()
    region_count.columns = ['Region', 'Count']

    tag_list = []
    for t in df['Tags']:
        tag_list.extend(t.keys())
    tag_count = pd.Series(tag_list).value_counts().reset_index()
    tag_count.columns = ['TagKey', 'Count']

    # Criação da planilha
    wb = Workbook()
    wb.remove(wb.active)

    # Aba 1: Inventário Completo
    ws1 = wb.create_sheet("Inventário Completo")
    for r in dataframe_to_rows(df[['ResourceARN', 'ResourceType', 'Region', 'TagsStr']], index=False, header=True):
        ws1.append(r)
    for cell in ws1["1:1"]:
        cell.font = Font(bold=True)

    # Aba 2: Top 10 Tipos
    ws2 = wb.create_sheet("Top 10 Tipos")
    for r in dataframe_to_rows(type_count.head(10), index=False, header=True):
        ws2.append(r)
    for cell in ws2["1:1"]:
        cell.font = Font(bold=True)

    chart1 = BarChart()
    chart1.title = "Top 10 Tipos de Recurso"
    max_row_types = len(type_count.head(10)) + 1
    chart1.add_data(Reference(ws2, min_col=2, min_row=1, max_row=max_row_types), titles_from_data=True)
    chart1.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=max_row_types))
    ws2.add_chart(chart1, "E2")

    # Aba 3: Recursos por Região
    ws3 = wb.create_sheet("Recursos por Região")
    for r in dataframe_to_rows(region_count, index=False, header=True):
        ws3.append(r)
    for cell in ws3["1:1"]:
        cell.font = Font(bold=True)

    chart2 = PieChart()
    chart2.title = "Distribuição por Região"
    chart2.add_data(Reference(ws3, min_col=2, min_row=1, max_row=len(region_count)+1), titles_from_data=True)
    chart2.set_categories(Reference(ws3, min_col=1, min_row=2, max_row=len(region_count)+1))
    ws3.add_chart(chart2, "E2")

    # Aba 4: Principais Tags
    ws4 = wb.create_sheet("Top Tags")
    for r in dataframe_to_rows(tag_count.head(10), index=False, header=True):
        ws4.append(r)
    for cell in ws4["1:1"]:
        cell.font = Font(bold=True)

    wb.save(output_path)

def main():
    print("🔍 Coletando recursos com tags...")
    df = get_tagged_resources()
    if df.empty:
        print("Nenhum recurso com tags encontrado.")
        return

    path = input("Digite o caminho e nome do arquivo para salvar (ex: C:/meus_relatorios/aws-inventory.xlsx): ").strip()
    if not path.endswith(".xlsx"):
        path += "/aws-tag-inventory.xlsx"

    export_to_excel(df, path)
    print(f"✅ Relatório salvo com sucesso em: {path}")

if __name__ == "__main__":
    main()
